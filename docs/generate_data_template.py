#!/usr/bin/env python3
"""Generate Excel template for client data collection."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Style definitions ──────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="2F5496")
NOTE_FONT = Font(name="Arial", italic=True, size=9, color="808080")
NORMAL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols, font=HEADER_FONT, fill=HEADER_FILL):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row, cols, is_required=True):
    fill = REQUIRED_FILL if is_required else OPTIONAL_FILL
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    # highlight the "required" column
    ws.cell(row=row, column=4).fill = fill


def add_section_title(ws, row, title, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    return row + 1


def add_note(ws, row, note, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = note
    cell.font = NOTE_FONT
    cell.alignment = WRAP
    return row + 1


def create_data_sheet(ws, title, headers, rows, notes=None, col_widths=None):
    """Create a formatted data sheet."""
    ws.title = title
    
    current_row = 1
    
    # Title
    current_row = add_section_title(ws, current_row, f"📋 {title}", len(headers))
    current_row += 1  # blank row
    
    # Notes if any
    if notes:
        for note in notes:
            current_row = add_note(ws, current_row, note, len(headers))
        current_row += 1
    
    # Headers
    for i, h in enumerate(headers, 1):
        ws.cell(row=current_row, column=i, value=h)
    style_header_row(ws, current_row, len(headers))
    current_row += 1
    
    # Data rows
    for row_data in rows:
        for i, val in enumerate(row_data, 1):
            ws.cell(row=current_row, column=i, value=val)
        is_req = row_data[3] == "✅ Bắt buộc" if len(row_data) > 3 else True
        style_data_row(ws, current_row, len(headers), is_req)
        current_row += 1
    
    # Column widths
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    
    # Freeze panes
    ws.freeze_panes = f"A{current_row - len(rows)}"
    
    return current_row


# ══════════════════════════════════════════════════════════════════
# SHEET 1: Hướng Dẫn (Instructions)
# ══════════════════════════════════════════════════════════════════
ws_intro = wb.active
ws_intro.title = "Hướng Dẫn"
ws_intro.sheet_properties.tabColor = "2F5496"

instructions = [
    ("📋 HƯỚNG DẪN CHUẨN BỊ DATA CHO WEBSITE", TITLE_FONT),
    ("", None),
    ("File này gồm các sheet tương ứng với từng loại dữ liệu cần chuẩn bị.", BOLD_FONT),
    ("Mỗi sheet có cột 'Bắt buộc' đánh dấu mức độ ưu tiên:", NORMAL_FONT),
    ("   • ✅ Bắt buộc  =  Phải có trước khi website đi vào hoạt động", NORMAL_FONT),
    ("   • ⬜ Tùy chọn  =  Có thể bổ sung sau, website vẫn hoạt động được", NORMAL_FONT),
    ("", None),
    ("CÁC SHEET TRONG FILE:", BOLD_FONT),
    ("   1. Thông Tin Trường — Thông tin cơ bản về trường/thương hiệu", NORMAL_FONT),
    ("   2. Khóa Học — Danh sách các khóa học/chương trình", NORMAL_FONT),
    ("   3. Giảng Viên — Thông tin từng giảng viên", NORMAL_FONT),
    ("   4. Lịch Học — Lịch cụ thể cho mỗi khóa", NORMAL_FONT),
    ("   5. Bài Viết — Blog, công thức, tin tức", NORMAL_FONT),
    ("   6. Testimonials — Nhận xét từ học viên", NORMAL_FONT),
    ("   7. Thanh Toán — Thông tin ngân hàng & VNPay", NORMAL_FONT),
    ("   8. Hình Ảnh — Danh sách ảnh cần chuẩn bị", NORMAL_FONT),
    ("   9. FAQ — Câu hỏi thường gặp", NORMAL_FONT),
    ("", None),
    ("CÁCH ĐIỀN:", BOLD_FONT),
    ("   • Điền vào cột 'Khách Hàng Điền Vào Đây' ở mỗi sheet", NORMAL_FONT),
    ("   • Ảnh: Đặt tên file rõ ràng và gửi kèm file Excel này", NORMAL_FONT),
    ("   • Nếu chưa có data cho mục nào, để trống, sẽ bổ sung sau", NORMAL_FONT),
    ("", None),
    ("Ô màu vàng nhạt = Bắt buộc  |  Ô màu xanh nhạt = Tùy chọn", BOLD_FONT),
]

for i, (text, font) in enumerate(instructions, 1):
    cell = ws_intro.cell(row=i, column=1, value=text)
    if font:
        cell.font = font
    cell.alignment = WRAP

ws_intro.column_dimensions["A"].width = 80


# ══════════════════════════════════════════════════════════════════
# SHEET 2: Thông Tin Trường
# ══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet()
headers = ["#", "Thông Tin", "Mô Tả / Ghi Chú", "Bắt Buộc", "Ví Dụ", "Khách Hàng Điền Vào Đây"]
rows = [
    [1, "Tên trường / thương hiệu", "Tên hiển thị trên website, logo", "✅ Bắt buộc", "Muka Cooking School", ""],
    [2, "Slogan / mô tả ngắn", "1-2 câu giới thiệu, hiển thị ở footer & header", "✅ Bắt buộc", "Trường dạy nấu ăn với hơn 10 năm kinh nghiệm đào tạo", ""],
    [3, "Địa chỉ", "Địa chỉ trường học", "✅ Bắt buộc", "123 Nguyễn Huệ, Quận 1, TP.HCM", ""],
    [4, "Số điện thoại", "SĐT liên hệ chính", "✅ Bắt buộc", "0901 234 567", ""],
    [5, "Email liên hệ", "Email nhận thư từ khách", "✅ Bắt buộc", "info@bakingschool.vn", ""],
    [6, "Website", "Tên miền website", "✅ Bắt buộc", "www.bakingschool.vn", ""],
    [7, "Giờ làm việc", "Giờ mở cửa / tiếp nhận", "✅ Bắt buộc", "T2-T7: 8:00 - 17:00", ""],
    [8, "Địa chỉ Google Maps", "Link Google Maps hoặc tọa độ", "✅ Bắt buộc", "https://maps.google.com/...", ""],
    [9, "Link Facebook", "Fanpage Facebook", "⬜ Tùy chọn", "https://facebook.com/bakingschool", ""],
    [10, "Link Instagram", "Trang Instagram", "⬜ Tùy chọn", "https://instagram.com/bakingschool", ""],
    [11, "Link YouTube", "Kênh YouTube", "⬜ Tùy chọn", "https://youtube.com/@bakingschool", ""],
    [12, "Link TikTok / khác", "Mạng xã hội khác", "⬜ Tùy chọn", "", ""],
]
create_data_sheet(ws2, "Thông Tin Trường", headers, rows,
    notes=["Đây là thông tin cơ bản về trường, hiển thị ở header, footer, trang About, và trang Contact."],
    col_widths=[5, 25, 45, 15, 40, 40])
ws2.sheet_properties.tabColor = "C00000"


# ══════════════════════════════════════════════════════════════════
# SHEET 3: Khóa Học
# ══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet()
headers = ["#", "Thông Tin", "Mô Tả / Ghi Chú", "Bắt Buộc", "Ví Dụ", "Khóa 1", "Khóa 2", "Khóa 3", "Khóa 4", "Khóa 5", "Khóa 6"]
rows = [
    [1, "Tên khóa học", "Tên hiển thị trên card và trang chi tiết", "✅ Bắt buộc", "Baking & Pastry Fundamentals", "", "", "", "", "", ""],
    [2, "Mô tả ngắn", "1-3 câu, hiển thị trên card khóa học", "✅ Bắt buộc", "Khóa học nền tảng về làm bánh, từ cơ bản đến nâng cao...", "", "", "", "", "", ""],
    [3, "Giá (VNĐ)", "Giá khóa học, số nguyên, đơn vị VNĐ", "✅ Bắt buộc", "550000", "", "", "", "", "", ""],
    [4, "Tên giảng viên phụ trách", "Tên GV dạy chính", "✅ Bắt buộc", "Nguyễn Thị B", "", "", "", "", "", ""],
    [5, "Ảnh thumbnail khóa học", "Tên file ảnh (gửi kèm). Đề xuất 800×600px", "✅ Bắt buộc", "khoa-baking.jpg", "", "", "", "", "", ""],
    [6, "Khóa nổi bật?", "Hiển thị trên slider trang chủ? (Có/Không)", "⬜ Tùy chọn", "Có", "", "", "", "", "", ""],
    [7, "Mục tiêu học (Learning Goals)", "Kỹ năng và % thành thạo.\nMỗi dòng 1 kỹ năng.\nFormat: Tên kỹ năng — %", "⬜ Tùy chọn", "Kỹ thuật nhào bột — 90\nTrang trí bánh — 75\nLàm kem — 80", "", "", "", "", "", ""],
    [8, "Khóa học bao gồm (Class Includes)", "Liệt kê những gì học viên nhận được.\nMỗi dòng 1 mục.", "⬜ Tùy chọn", "Nguyên liệu thực hành\nGiáo trình in sẵn\nChứng chỉ hoàn thành\nVideo replay 30 ngày", "", "", "", "", "", ""],
    [9, "Giáo trình (Curriculum)", "Chia theo module.\nFormat: Tên module — Nội dung", "⬜ Tùy chọn", "Module 1: Cơ bản — Nguyên liệu, dụng cụ\nModule 2: Bánh mì — Các loại bánh mì\nModule 3: Bánh ngọt — Cake, tart", "", "", "", "", "", ""],
    [10, "Video bài giảng (Premium)", "Link YouTube embed.\nChỉ xem được sau khi mua.\nFormat: Tên — Link", "⬜ Tùy chọn", "Bài 1: Nhào bột — https://youtube.com/embed/xxx\nBài 2: Nướng bánh — https://youtube.com/embed/yyy", "", "", "", "", "", ""],
    [11, "Tài liệu tải về (Premium)", "File PDF, tài liệu.\nFormat: Tên — Link download", "⬜ Tùy chọn", "Công thức bánh mì — link\nBảng quy đổi nguyên liệu — link", "", "", "", "", "", ""],
    [12, "Hướng dẫn chi tiết (Premium)", "Text dài, hướng dẫn bổ sung.\nChỉ xem được sau khi mua.", "⬜ Tùy chọn", "Hướng dẫn chi tiết cách ủ bột qua đêm...", "", "", "", "", "", ""],
]
create_data_sheet(ws3, "Khóa Học", headers, rows,
    notes=[
        "Mỗi cột (Khóa 1, Khóa 2...) tương ứng với 1 khóa học. Điền thông tin vào từng cột.",
        "Nếu nhiều hơn 6 khóa, thêm cột hoặc tạo thêm dòng bên dưới.",
        "Premium Content (dòng 10-12): Nội dung chỉ hiển thị cho học viên đã mua khóa."
    ],
    col_widths=[5, 30, 45, 15, 45, 30, 30, 30, 30, 30, 30])
ws3.sheet_properties.tabColor = "C00000"


# ══════════════════════════════════════════════════════════════════
# SHEET 4: Giảng Viên
# ══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet()
headers = ["#", "Thông Tin", "Mô Tả / Ghi Chú", "Bắt Buộc", "Ví Dụ", "GV 1", "GV 2", "GV 3", "GV 4", "GV 5", "GV 6"]
rows = [
    [1, "Họ tên", "Tên đầy đủ giảng viên", "✅ Bắt buộc", "Nguyễn Thị Bảo Trân", "", "", "", "", "", ""],
    [2, "Chức danh", "Vai trò/title", "✅ Bắt buộc", "Master Chef / Pastry Chef / Assistant", "", "", "", "", "", ""],
    [3, "Ảnh chân dung", "Tên file ảnh (gửi kèm).\nĐề xuất: 600×700px, nền sạch", "✅ Bắt buộc", "gv-tran.jpg", "", "", "", "", "", ""],
    [4, "Giới thiệu ngắn (bio)", "2-3 câu giới thiệu", "⬜ Tùy chọn", "Với hơn 10 năm kinh nghiệm trong nghề bánh, chị Trân đã đào tạo hơn 500 học viên...", "", "", "", "", "", ""],
    [5, "Tiểu sử chi tiết (biography)", "Nhiều đoạn văn.\nQuá trình học tập, thành tựu.", "⬜ Tùy chọn", "Tốt nghiệp Le Cordon Bleu Paris năm 2010...\nTừng làm việc tại...", "", "", "", "", "", ""],
    [6, "Điểm nổi bật (highlights)", "Các thành tích.\nPhân cách bằng dấu |", "⬜ Tùy chọn", "Giải nhất cuộc thi X | 15 năm kinh nghiệm | Đào tạo 500+ học viên", "", "", "", "", "", ""],
    [7, "Kỹ năng (skills)", "Kỹ năng và % thành thạo.\nFormat: Tên — %", "⬜ Tùy chọn", "Làm bánh Pháp — 95\nTrang trí — 88\nSourdough — 92", "", "", "", "", "", ""],
    [8, "Link Facebook", "", "⬜ Tùy chọn", "https://facebook.com/...", "", "", "", "", "", ""],
    [9, "Link Twitter/X", "", "⬜ Tùy chọn", "", "", "", "", "", "", ""],
    [10, "Link LinkedIn / khác", "", "⬜ Tùy chọn", "", "", "", "", "", "", ""],
]
create_data_sheet(ws4, "Giảng Viên", headers, rows,
    notes=["Mỗi cột (GV 1, GV 2...) tương ứng 1 giảng viên. Đề xuất: 3-6 giảng viên."],
    col_widths=[5, 28, 45, 15, 45, 30, 30, 30, 30, 30, 30])
ws4.sheet_properties.tabColor = "C00000"


# ══════════════════════════════════════════════════════════════════
# SHEET 5: Lịch Học
# ══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet()
headers = ["#", "Khóa Học", "Ngày Bắt Đầu", "Ngày Kết Thúc", "Hạn Đăng Ký", "Ngày Học", "Khung Giờ", "GV Thay Thế (nếu có)"]
sample_rows = [
    [1, "Baking & Pastry", "01/05/2026", "15/06/2026", "25/04/2026", "Thứ 2, Thứ 4", "09:00 - 11:30", ""],
    [2, "Baking & Pastry", "01/07/2026", "15/08/2026", "25/06/2026", "Thứ 3, Thứ 5", "14:00 - 16:30", ""],
    [3, "(Điền tên khóa)", "", "", "", "", "", ""],
    [4, "", "", "", "", "", "", ""],
    [5, "", "", "", "", "", "", ""],
    [6, "", "", "", "", "", "", ""],
    [7, "", "", "", "", "", "", ""],
    [8, "", "", "", "", "", "", ""],
    [9, "", "", "", "", "", "", ""],
    [10, "", "", "", "", "", "", ""],
    [11, "", "", "", "", "", "", ""],
    [12, "", "", "", "", "", "", ""],
]

ws5.title = "Lịch Học"
ws5.sheet_properties.tabColor = "C00000"

r = 1
r = add_section_title(ws5, r, "📅 Lịch Học (Class Sessions)", 8)
r = add_note(ws5, r, "Mỗi khóa học có thể có nhiều lịch/đợt. Mỗi dòng = 1 lịch.", 8)
r = add_note(ws5, r, "Tất cả các cột đều bắt buộc (trừ GV Thay Thế). Thêm dòng nếu cần.", 8)
r += 1

for i, h in enumerate(headers, 1):
    ws5.cell(row=r, column=i, value=h)
style_header_row(ws5, r, len(headers))
r += 1

for row_data in sample_rows:
    for i, val in enumerate(row_data, 1):
        cell = ws5.cell(row=r, column=i, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    r += 1

for i, w in enumerate([5, 25, 18, 18, 18, 20, 20, 25], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════
# SHEET 6: Bài Viết
# ══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet()
headers = ["#", "Thông Tin", "Mô Tả / Ghi Chú", "Bắt Buộc", "Ví Dụ", "Bài 1", "Bài 2", "Bài 3"]
rows = [
    [1, "Tiêu đề bài viết", "", "✅ Bắt buộc", "Bí Quyết Làm Bánh Mì Sourdough Hoàn Hảo", "", "", ""],
    [2, "Loại bài", "BLOG hoặc RECIPE", "✅ Bắt buộc", "RECIPE", "", "", ""],
    [3, "Danh mục", "Phân loại bài viết", "✅ Bắt buộc", "Recipes / Classes / Tips / News", "", "", ""],
    [4, "Mô tả ngắn", "1-2 câu hiển thị ở card preview", "✅ Bắt buộc", "Hướng dẫn chi tiết cách làm bánh mì sourdough từ men tự nhiên", "", "", ""],
    [5, "Nội dung đầy đủ", "Bài viết đầy đủ.\nCó thể gửi file Word riêng.", "✅ Bắt buộc", "(Gửi file Word kèm theo)", "", "", ""],
    [6, "Ảnh thumbnail", "Tên file ảnh.\nĐề xuất: 800×500px", "✅ Bắt buộc", "blog-sourdough.jpg", "", "", ""],
    [7, "Tên tác giả", "Mặc định: Admin", "⬜ Tùy chọn", "Admin", "", "", ""],
    [8, "Ngày đăng", "Nếu trống sẽ dùng ngày nhập liệu", "⬜ Tùy chọn", "15/04/2026", "", "", ""],
]
create_data_sheet(ws6, "Bài Viết", headers, rows,
    notes=["Đề xuất: chuẩn bị 3-6 bài viết ban đầu. Có thể thêm sau qua trang Admin."],
    col_widths=[5, 25, 40, 15, 45, 35, 35, 35])
ws6.sheet_properties.tabColor = "ED7D31"


# ══════════════════════════════════════════════════════════════════
# SHEET 7: Testimonials
# ══════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet()
headers = ["#", "Tên Người Nhận Xét", "Vai Trò", "Trích Dẫn Ngắn (1 câu)", "Nhận Xét Đầy Đủ (3-5 câu)"]

ws7.title = "Testimonials"
ws7.sheet_properties.tabColor = "ED7D31"

r = 1
r = add_section_title(ws7, r, "💬 Nhận Xét Từ Học Viên (Testimonials)", 5)
r = add_note(ws7, r, "Đề xuất: 3-5 nhận xét. Tất cả các cột đều bắt buộc.", 5)
r += 1

for i, h in enumerate(headers, 1):
    ws7.cell(row=r, column=i, value=h)
style_header_row(ws7, r, len(headers))
r += 1

samples = [
    [1, "Lê Thị D", "Cựu học viên / Pastry Chef", "Khóa học đã thay đổi sự nghiệp của tôi!", "Tôi đã từ một người hoàn toàn không biết gì về làm bánh trở thành pastry chef chuyên nghiệp. Giảng viên rất tận tình và nội dung thực tế."],
    [2, "(Điền tên)", "", "", ""],
    [3, "", "", "", ""],
    [4, "", "", "", ""],
    [5, "", "", "", ""],
]
for row_data in samples:
    for i, val in enumerate(row_data, 1):
        cell = ws7.cell(row=r, column=i, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    r += 1

for i, w in enumerate([5, 25, 25, 40, 60], 1):
    ws7.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════
# SHEET 8: Thanh Toán
# ══════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet()
headers = ["#", "Thông Tin", "Mô Tả / Ghi Chú", "Bắt Buộc", "Khách Hàng Điền Vào Đây"]
rows = [
    [1, "Tên ngân hàng", "VD: Vietcombank, Techcombank, MB Bank...", "✅ Bắt buộc", ""],
    [2, "Số tài khoản", "Số TK nhận thanh toán từ học viên", "✅ Bắt buộc", ""],
    [3, "Tên chủ tài khoản", "Viết IN HOA, đúng như trên ngân hàng", "✅ Bắt buộc", ""],
    [4, "Ảnh QR chuyển khoản", "Ảnh QR code ngân hàng (gửi file kèm)", "⬜ Tùy chọn", ""],
    [5, "Ghi chú chuyển khoản mẫu", "Mẫu nội dung CK cho học viên", "⬜ Tùy chọn", ""],
]

# VNPay section
vnpay_rows = [
    [6, "───── VNPAY (nếu tích hợp) ─────", "", "", ""],
    [7, "VNPay TMN Code", "Mã terminal do VNPay cấp", "✅ Bắt buộc *", ""],
    [8, "VNPay Hash Secret", "Secret key do VNPay cấp (BẢO MẬT)", "✅ Bắt buộc *", ""],
    [9, "Sandbox hay Production?", "Đang test hay lên thật?", "✅ Bắt buộc *", ""],
    [10, "* Chỉ bắt buộc nếu tích hợp VNPay", "", "", ""],
]

all_rows = rows + vnpay_rows
create_data_sheet(ws8, "Thanh Toán", headers, all_rows,
    notes=["Thông tin ngân hàng để hiển thị cho học viên khi thanh toán chuyển khoản."],
    col_widths=[5, 30, 50, 18, 40])
ws8.sheet_properties.tabColor = "C00000"


# ══════════════════════════════════════════════════════════════════
# SHEET 9: Hình Ảnh
# ══════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet()
headers = ["#", "Loại Ảnh", "Số Lượng", "Kích Thước Đề Xuất", "Bắt Buộc", "Ghi Chú", "Tên File"]
rows = [
    [1, "Logo", "1", "PNG trong suốt, ~200×60px", "✅ Bắt buộc", "File PNG nền trong suốt", ""],
    [2, "Favicon", "1", "32×32px hoặc 64×64px", "✅ Bắt buộc", "Icon nhỏ hiển thị trên tab trình duyệt", ""],
    [3, "Ảnh slider trang chủ", "2-3 tấm", "1920×800px trở lên", "✅ Bắt buộc", "Ảnh banner toàn màn hình, chất lượng cao", ""],
    [4, "Ảnh thumbnail khóa học", "1 ảnh/khóa", "800×600px", "✅ Bắt buộc", "Ảnh đại diện từng khóa", ""],
    [5, "Ảnh giảng viên (chân dung)", "1 ảnh/người", "600×700px, tỷ lệ đứng", "✅ Bắt buộc", "Ảnh chân dung, nền sạch", ""],
    [6, "Ảnh bài viết blog", "1 ảnh/bài", "800×500px", "✅ Bắt buộc", "Thumbnail cho mỗi bài viết", ""],
    [7, "Ảnh/Video giới thiệu (About)", "1", "Ảnh hoặc link YouTube", "⬜ Tùy chọn", "Video giới thiệu về trường", ""],
    [8, "Ảnh gallery hoạt động", "4-10 tấm", "Tùy ý", "⬜ Tùy chọn", "Ảnh hoạt động, lớp học, sản phẩm", ""],
    [9, "Ảnh QR thanh toán", "1", "Tùy ý", "⬜ Tùy chọn", "QR code ngân hàng", ""],
]
create_data_sheet(ws9, "Hình Ảnh", headers, rows,
    notes=[
        "Gửi tất cả ảnh trong 1 folder kèm file Excel này.",
        "Đặt tên file rõ ràng, ví dụ: logo.png, slider-01.jpg, khoa-baking.jpg, gv-tran.jpg",
        "Ưu tiên ảnh chất lượng cao, chưa bị nén. Định dạng: JPG hoặc PNG."
    ],
    col_widths=[5, 28, 15, 28, 15, 45, 30])
ws9.sheet_properties.tabColor = "ED7D31"


# ══════════════════════════════════════════════════════════════════
# SHEET 10: FAQ
# ══════════════════════════════════════════════════════════════════
ws10 = wb.create_sheet()
headers = ["#", "Câu Hỏi", "Câu Trả Lời"]

ws10.title = "FAQ"
ws10.sheet_properties.tabColor = "ED7D31"

r = 1
r = add_section_title(ws10, r, "❓ Câu Hỏi Thường Gặp (FAQ)", 3)
r = add_note(ws10, r, "Đề xuất: 4-6 câu hỏi phổ biến. Hiển thị ở trang chủ.", 3)
r += 1

for i, h in enumerate(headers, 1):
    ws10.cell(row=r, column=i, value=h)
style_header_row(ws10, r, len(headers))
r += 1

faqs = [
    [1, "Khóa học kéo dài bao lâu?", "(Điền câu trả lời)"],
    [2, "Cần chuẩn bị gì trước khi tham gia?", ""],
    [3, "Chính sách hoàn tiền như thế nào?", ""],
    [4, "Có cấp chứng chỉ sau khóa học không?", ""],
    [5, "Học phí đã bao gồm nguyên liệu chưa?", ""],
    [6, "Có hỗ trợ học lại không?", ""],
]
for row_data in faqs:
    for i, val in enumerate(row_data, 1):
        cell = ws10.cell(row=r, column=i, value=val)
        cell.font = NORMAL_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    r += 1

for i, w in enumerate([5, 45, 60], 1):
    ws10.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════
output_path = "/Users/hoangkien/NLV/baking/docs/DATA_TEMPLATE_CHO_KHACH_HANG.xlsx"
wb.save(output_path)
print(f"✅ File Excel đã được tạo: {output_path}")
